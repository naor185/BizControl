"""
Expense management routes for the SaaS Business Management module.
All queries are strictly filtered by studio_id for multi-tenant data isolation.
"""
from __future__ import annotations

import logging
import os
import re
import uuid

_log = logging.getLogger(__name__)
from datetime import date
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.deps import require_studio_ctx, AuthContext
from app.db.deps import get_db
from app.repositories.expense_repository import ExpenseRepository
from app.schemas.expense import ExpenseCreate, ExpenseResponse, ExpenseUpdate, ExpenseSummary
from app.services.ai_invoice_service import AIInvoiceService
from datetime import datetime

router = APIRouter(prefix="/expenses", tags=["Expenses"])


def _save_receipt_image(image_bytes: bytes, content_type: str, studio_id: str, db=None) -> str | None:
    """Save receipt image — Cloudinary if configured, otherwise local uploads/."""
    ext = (content_type or "image/jpeg").split("/")[-1].replace("jpeg", "jpg")
    public_id = f"receipt_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}"
    try:
        from app.api.upload_routes import _cloudinary_upload
        cloud_url = _cloudinary_upload(image_bytes, folder=f"receipts/{studio_id}", public_id=public_id, db=db)
        if cloud_url:
            return cloud_url
    except Exception as e:
        _log.debug("Cloudinary not available: %s", e)
    # Fall back to local filesystem
    try:
        upload_dir = os.path.join("uploads", "receipts", studio_id)
        os.makedirs(upload_dir, exist_ok=True)
        fname = f"{public_id}.{ext}"
        with open(os.path.join(upload_dir, fname), "wb") as fh:
            fh.write(image_bytes)
        return f"/uploads/receipts/{studio_id}/{fname}"
    except Exception as e:
        _log.warning("Could not save receipt image: %s", e)
        return None


def _delete_receipt_file(url: str | None, db=None) -> None:
    """Best-effort delete of a receipt image — Cloudinary asset or local file."""
    if not url:
        return
    if url.startswith("http"):
        cloud_name = os.getenv("CLOUDINARY_CLOUD_NAME", "").strip()
        if not cloud_name:
            return
        m = re.search(r"/upload/(?:v\d+/)?(.+?)\.\w+(?:\?.*)?$", url)
        if not m:
            return
        public_id = m.group(1)
        try:
            import cloudinary
            import cloudinary.uploader
            cloudinary.config(
                cloud_name=cloud_name,
                api_key=os.getenv("CLOUDINARY_API_KEY", ""),
                api_secret=os.getenv("CLOUDINARY_API_SECRET", ""),
            )
            cloudinary.uploader.destroy(public_id, resource_type="image")
        except Exception as e:
            _log.warning("Could not delete Cloudinary receipt %s: %s", public_id, e)
            if db is not None:
                from app.services.integration_alerts import alert_integration_failure
                alert_integration_failure(db, "Cloudinary (מחיקת תמונות)", str(e))
    else:
        file_path = os.path.normpath(url.lstrip("/"))
        uploads_root = os.path.normpath("uploads")
        if os.path.abspath(file_path).startswith(os.path.abspath(uploads_root)) and os.path.exists(file_path):
            try:
                os.remove(file_path)
            except OSError as e:
                _log.warning("Could not delete local receipt %s: %s", file_path, e)


def _measure_receipt_size(url: str | None) -> int | None:
    """Best-effort byte size lookup for a legacy receipt that predates file_size_bytes tracking."""
    if not url:
        return None
    if url.startswith("http"):
        try:
            import urllib.request
            req = urllib.request.Request(url, method="HEAD")
            with urllib.request.urlopen(req, timeout=8) as resp:
                length = resp.headers.get("Content-Length")
                return int(length) if length else None
        except Exception:
            return None
    file_path = os.path.normpath(url.lstrip("/"))
    try:
        return os.path.getsize(file_path)
    except OSError:
        return None


def _reset_scan_quota_if_new_month(db: Session, studio) -> None:
    """Mirror the monthly quota rollover so /scan and /storage/usage never disagree.
    Commits immediately on rollover, matching the previous /scan-only behavior."""
    from datetime import datetime
    current_month = datetime.utcnow().strftime("%Y-%m")
    if studio.invoice_scan_reset_month != current_month:
        studio.invoice_scan_used = 0
        studio.invoice_scan_prompt_tokens = 0
        studio.invoice_scan_completion_tokens = 0
        studio.invoice_scan_cost_usd = Decimal("0")
        studio.invoice_scan_reset_month = current_month
        db.commit()


def get_expense_repo(db: Session = Depends(get_db)) -> ExpenseRepository:
    return ExpenseRepository(db)


# ── List / Filter ───────────────────────────────────────────────────────────
@router.get("", response_model=list[ExpenseResponse])
def list_expenses(
    month: Optional[int] = Query(None, ge=1, le=12, description="Filter by month (1-12)"),
    year: Optional[int] = Query(None, ge=2000, le=2100, description="Filter by year"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    ctx: AuthContext = Depends(require_studio_ctx),
    repo: ExpenseRepository = Depends(get_expense_repo),
):
    return repo.get_multi(
        studio_id=ctx.studio_id,
        skip=skip,
        limit=limit,
        month=month,
        year=year,
    )


# ── Monthly Summary (Dashboard Cards) ────────────────────────────────────────
@router.get("/summary", response_model=ExpenseSummary)
def expense_summary(
    month: int = Query(..., ge=1, le=12, description="Month to summarize"),
    year: int = Query(..., ge=2000, le=2100, description="Year to summarize"),
    ctx: AuthContext = Depends(require_studio_ctx),
    repo: ExpenseRepository = Depends(get_expense_repo),
):
    total, vat, count = repo.get_monthly_summary(
        studio_id=ctx.studio_id,
        month=month,
        year=year,
    )
    return ExpenseSummary(total_expenses=total, total_vat=vat, invoice_count=count)


# ── Create (Manual Entry) ────────────────────────────────────────────────────
@router.post("", response_model=ExpenseResponse, status_code=status.HTTP_201_CREATED)
def create_expense(
    payload: ExpenseCreate,
    ctx: AuthContext = Depends(require_studio_ctx),
    repo: ExpenseRepository = Depends(get_expense_repo),
):
    return repo.create(studio_id=ctx.studio_id, expense_in=payload)


# ── Duplicate check (warn before saving, never blocks) ────────────────────────
@router.get("/check-duplicate")
def check_duplicate_expense(
    supplier: str = Query(...),
    date: date = Query(...),
    amount: Decimal = Query(...),
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    from sqlalchemy import func

    supplier_norm = supplier.strip().lower()
    if not supplier_norm:
        return {"is_duplicate": False, "existing_id": None}

    existing = db.query(ExpenseModel).filter(
        ExpenseModel.studio_id == ctx.studio_id,
        ExpenseModel.expense_date == date,
        ExpenseModel.amount == amount,
        func.lower(func.trim(func.coalesce(ExpenseModel.supplier_name, ExpenseModel.title))) == supplier_norm,
    ).first()

    return {"is_duplicate": existing is not None, "existing_id": str(existing.id) if existing else None}


# ── AI Invoice Scan ──────────────────────────────────────────────────────────
@router.post("/scan", status_code=status.HTTP_200_OK)
async def scan_invoice(
    file: UploadFile = File(..., description="Invoice image (JPG, PNG, WEBP)"),
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    """
    Upload an invoice image and extract structured data via Document AI.
    Requires feature flag 'invoice_ai_scan' enabled and quota not exceeded.
    """
    from app.models.studio import Studio
    from app.models.studio_feature import StudioFeature

    # Check feature flag
    feature = db.query(StudioFeature).filter_by(
        studio_id=ctx.studio_id, feature="invoice_ai_scan"
    ).first()
    if not feature or not feature.is_enabled:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="סריקת חשבוניות AI אינה מופעלת לסטודיו זה. צרו קשר עם התמיכה.",
        )

    # Check quota + reset monthly
    studio = db.query(Studio).filter_by(id=ctx.studio_id).first()
    if studio:
        _reset_scan_quota_if_new_month(db, studio)
        if studio.invoice_scan_quota > 0 and studio.invoice_scan_used >= studio.invoice_scan_quota:
            raise HTTPException(
                status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                detail=f"חרגתם ממכסת הסריקות החודשית ({studio.invoice_scan_quota} סריקות). פנו לתמיכה להגדלת המכסה.",
            )

    allowed_types = {"image/jpeg", "image/png", "image/webp", "image/heic", "image/heif"}
    content_type = (file.content_type or "").split(";")[0].strip().lower()
    if content_type not in allowed_types:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"סוג קובץ לא נתמך '{content_type}'. יש להעלות JPG, PNG, WEBP.",
        )

    adc_json = os.getenv("GOOGLE_ADC_JSON", "").strip()
    openai_key = os.getenv("OPENAI_API_KEY", "").strip()
    gemini_key = os.getenv("GEMINI_API_KEY", "").strip()
    project_id = os.getenv("DOCUMENT_AI_PROJECT_ID", "").strip()
    processor_id = os.getenv("DOCUMENT_AI_PROCESSOR_ID", "").strip()
    has_vision_key = (
        (adc_json and project_id and processor_id) or
        (openai_key and openai_key.startswith("sk-")) or
        (gemini_key and gemini_key.startswith("AIza")) or
        (openai_key and openai_key.startswith("AIza"))
    )
    if not has_vision_key:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="סריקת חשבוניות דורשת GEMINI_API_KEY, OPENAI_API_KEY (sk-...) או Google Document AI.",
        )

    image_bytes = await file.read()

    try:
        service = AIInvoiceService()
        result = service.parse_invoice_from_bytes(image_bytes, content_type=file.content_type)
    except Exception as e:
        from app.services.integration_alerts import alert_integration_failure
        alert_integration_failure(db, "סריקת חשבוניות AI (Gemini/OpenAI/Document AI)", str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"AI parsing failed: {str(e)}",
        )

    # Save receipt image to disk
    receipt_url = _save_receipt_image(image_bytes, content_type, str(ctx.studio_id), db=db)

    # Increment usage counter + AI cost/token tracking
    if studio:
        studio.invoice_scan_used = (studio.invoice_scan_used or 0) + 1
        if result.prompt_tokens is not None:
            studio.invoice_scan_prompt_tokens = (studio.invoice_scan_prompt_tokens or 0) + result.prompt_tokens
        if result.completion_tokens is not None:
            studio.invoice_scan_completion_tokens = (studio.invoice_scan_completion_tokens or 0) + result.completion_tokens
        from app.services.ai_invoice_service import _estimate_cost_usd
        cost = _estimate_cost_usd(result.model_used, result.prompt_tokens, result.completion_tokens)
        if cost is not None:
            studio.invoice_scan_cost_usd = (studio.invoice_scan_cost_usd or Decimal("0")) + cost
        db.commit()

    return {
        "business_name": result.business_name,
        "invoice_number": result.invoice_number,
        "total_amount": float(result.total_amount) if result.total_amount else None,
        "vat_amount": float(result.vat_amount) if result.vat_amount else None,
        "pretax_amount": float(result.pretax_amount) if result.pretax_amount else None,
        "invoice_date": result.invoice_date.isoformat() if result.invoice_date else None,
        "payment_method": result.payment_method,
        "receipt_url": receipt_url,
        "receipt_size_bytes": len(image_bytes),
        "ai_provider": service._provider,
    }


# ── Mark as sent to accountant ───────────────────────────────────────────────
@router.post("/{expense_id}/mark-sent", response_model=ExpenseResponse)
def mark_sent_to_accountant(
    expense_id: uuid.UUID,
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    exp = db.query(ExpenseModel).filter_by(id=expense_id, studio_id=ctx.studio_id).first()
    if not exp:
        raise HTTPException(status_code=404, detail="Expense not found")
    exp.sent_to_accountant = True
    exp.sent_to_accountant_at = datetime.utcnow()
    db.commit()
    db.refresh(exp)
    return exp


@router.post("/{expense_id}/unmark-sent", response_model=ExpenseResponse)
def unmark_sent_to_accountant(
    expense_id: uuid.UUID,
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    exp = db.query(ExpenseModel).filter_by(id=expense_id, studio_id=ctx.studio_id).first()
    if not exp:
        raise HTTPException(status_code=404, detail="Expense not found")
    exp.sent_to_accountant = False
    exp.sent_to_accountant_at = None
    db.commit()
    db.refresh(exp)
    return exp


# ── Mark ALL in month as sent ────────────────────────────────────────────────
@router.post("/mark-month-sent", status_code=status.HTTP_200_OK)
def mark_month_sent(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2000),
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    from sqlalchemy import and_, extract
    now = datetime.utcnow()
    db.query(ExpenseModel).filter(
        ExpenseModel.studio_id == ctx.studio_id,
        ExpenseModel.sent_to_accountant == False,
        extract("month", ExpenseModel.expense_date) == month,
        extract("year", ExpenseModel.expense_date) == year,
    ).update({"sent_to_accountant": True, "sent_to_accountant_at": now}, synchronize_session=False)
    db.commit()
    return {"ok": True}


# ── Send a date range of expenses to accountant by email ──────────────────────
@router.post("/send-to-accountant", status_code=status.HTTP_200_OK)
def send_expenses_to_accountant(
    body: dict,
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    """Email expenses in a date range to the studio's accountant via the central
    Email Center (app/services/email_center.py) — studios no longer configure
    their own Resend credentials."""
    from sqlalchemy import text as sa_text
    from app.models.studio import Studio
    from app.models.expense import Expense as ExpenseModel

    date_from = body.get("date_from")
    date_to = body.get("date_to")
    if not date_from or not date_to:
        raise HTTPException(status_code=400, detail="חסרים תאריכים")

    inv_settings = db.execute(
        sa_text("SELECT accountant_email FROM invoice_settings WHERE studio_id = :sid"),
        {"sid": str(ctx.studio_id)},
    ).fetchone()
    accountant_email = inv_settings[0] if inv_settings else None
    if not accountant_email:
        raise HTTPException(status_code=400, detail="לא הוגדר מייל רואה חשבון. הגדר אותו בטאב ההגדרות בעמוד החשבוניות.")

    studio = db.get(Studio, ctx.studio_id)
    biz_name = studio.name if studio else ""

    expenses = db.query(ExpenseModel).filter(
        ExpenseModel.studio_id == ctx.studio_id,
        ExpenseModel.expense_date >= date_from,
        ExpenseModel.expense_date <= date_to,
    ).order_by(ExpenseModel.expense_date).all()
    if not expenses:
        raise HTTPException(status_code=400, detail="אין הוצאות בטווח התאריכים שנבחר")

    api_base = os.getenv("API_BASE_URL", "").rstrip("/")

    def _receipt_link(url: str | None) -> str:
        if not url:
            return ""
        if url.startswith("http"):
            full = url
        elif api_base:
            full = f"{api_base}{url}"
        else:
            # No backend base URL configured — a bare relative path like
            # "/uploads/..." has no page to resolve against inside an email
            # client, and gets mangled into a broken "http:///..." link.
            # Omit the link rather than send something that can't work.
            return "—"
        return f'<a href="{full}" style="color:#4f46e5">תמונת קבלה</a>'

    method_labels = {
        "אשראי": "כרטיס אשראי", "מזומן": "מזומן", "ביט/פייבוקס": "Bit / PayBox",
        "העברה בנקאית": "העברה בנקאית", "צ'ק": "צ'ק",
    }

    rows_html = ""
    total_sum = Decimal("0")
    vat_sum = Decimal("0")
    for exp in expenses:
        total_sum += exp.amount or Decimal("0")
        vat_sum += exp.vat_amount or Decimal("0")
        expense_date_str = exp.expense_date.strftime("%d/%m/%Y") if exp.expense_date else ""
        rows_html += f"""
        <tr>
          <td style="padding:8px 12px;border-bottom:1px solid #f0f0f0">{exp.supplier_name or exp.title}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #f0f0f0">{exp.category or '—'}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #f0f0f0;text-align:center">{expense_date_str}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #f0f0f0;text-align:center">{method_labels.get(exp.payment_method or '', exp.payment_method or '—')}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #f0f0f0;text-align:left;font-weight:700">₪{exp.amount:.2f}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #f0f0f0;text-align:left">₪{(exp.vat_amount or Decimal('0')):.2f}</td>
          <td style="padding:8px 12px;border-bottom:1px solid #f0f0f0;text-align:center">{_receipt_link(exp.receipt_url)}</td>
        </tr>"""

    html = f"""
    <html dir="rtl" lang="he"><head><meta charset="UTF-8">
    <style>body{{font-family:Arial,sans-serif;direction:rtl;color:#1a1a2e}}
    table{{width:100%;border-collapse:collapse}}th{{background:#1a1a2e;color:#fff;padding:10px 12px;font-size:13px}}</style>
    </head><body>
    <div style="max-width:800px;margin:0 auto;padding:20px">
      <h2 style="color:#1a1a2e;margin-bottom:4px">דוח הוצאות {date_from} עד {date_to} — {biz_name}</h2>
      <p style="color:#64748b;font-size:13px;margin-bottom:20px">סה"כ {len(expenses)} הוצאות</p>
      <table>
        <thead><tr>
          <th>ספק</th><th>קטגוריה</th><th>תאריך</th><th>אמצעי תשלום</th>
          <th style="text-align:left">סכום</th><th style="text-align:left">מע"מ</th><th>קבלה</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
        <tfoot><tr style="background:#f8fafc;font-weight:700">
          <td colspan="4" style="padding:10px 12px;text-align:right">סה"כ</td>
          <td style="padding:10px 12px;text-align:left">₪{total_sum:.2f}</td>
          <td style="padding:10px 12px;text-align:left">₪{vat_sum:.2f}</td>
          <td></td>
        </tr></tfoot>
      </table>
      <p style="color:#94a3b8;font-size:11px;margin-top:20px">נשלח ממערכת BizControl</p>
    </div></body></html>"""

    from app.services.email_center import send_email as ec_send_email
    ok = ec_send_email(
        db,
        to_email=accountant_email,
        subject=f"דוח הוצאות {date_from} עד {date_to} | {biz_name}",
        html_content=html,
        from_name=biz_name or "BizControl",
        studio_id=str(ctx.studio_id),
        template_key="expenses_report",
        email_type="invoice",
    )
    if not ok:
        raise HTTPException(
            status_code=502,
            detail="שליחת המייל נכשלה. בדוק את הגדרות מרכז המייל (סופר-אדמין → מרכז מייל → לוגים).",
        )

    now = datetime.utcnow()
    db.query(ExpenseModel).filter(
        ExpenseModel.studio_id == ctx.studio_id,
        ExpenseModel.expense_date >= date_from,
        ExpenseModel.expense_date <= date_to,
    ).update({"sent_to_accountant": True, "sent_to_accountant_at": now}, synchronize_session=False)
    db.commit()

    return {"ok": True, "sent_count": len(expenses)}


# ── Export month's receipt images as a ZIP ────────────────────────────────────
@router.get("/export/receipts-zip")
def export_receipts_zip(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2000),
    ctx: AuthContext = Depends(require_studio_ctx),
    repo: ExpenseRepository = Depends(get_expense_repo),
):
    """Bundle this month's receipt images into a single ZIP for manual handling
    (e.g. attaching to an email yourself) — the automated accountant email only
    links to images (see /send-to-accountant) since attaching many files risks
    exceeding email provider size limits."""
    import io
    import zipfile
    import requests

    expenses = repo.get_multi(studio_id=ctx.studio_id, month=month, year=year, limit=1000)
    expenses_with_receipts = [e for e in expenses if e.receipt_url]
    if not expenses_with_receipts:
        raise HTTPException(status_code=400, detail="אין קבלות עם תמונה לחודש זה")

    buf = io.BytesIO()
    used_names: set[str] = set()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for exp in expenses_with_receipts:
            url = exp.receipt_url
            try:
                if url.startswith("http"):
                    resp = requests.get(url, timeout=15)
                    resp.raise_for_status()
                    content = resp.content
                    ext = url.split("?")[0].rsplit(".", 1)[-1].lower()
                    ext = ext if ext in ("jpg", "jpeg", "png", "webp", "heic") else "jpg"
                else:
                    file_path = os.path.normpath(url.lstrip("/"))
                    with open(file_path, "rb") as fh:
                        content = fh.read()
                    ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else "jpg"
            except Exception as e:
                _log.warning("Could not fetch receipt %s for zip export: %s", url, e)
                continue

            date_str = exp.expense_date.strftime("%Y-%m-%d") if exp.expense_date else "unknown"
            supplier = re.sub(r"[^\w\-א-ת ]", "", exp.supplier_name or exp.title)[:40].strip() or "קבלה"
            name = f"{date_str}_{supplier}.{ext}"
            counter = 2
            while name in used_names:
                name = f"{date_str}_{supplier}_{counter}.{ext}"
                counter += 1
            used_names.add(name)
            zf.writestr(name, content)

    if not used_names:
        raise HTTPException(status_code=502, detail="לא הצלחנו להוריד אף תמונת קבלה")

    buf.seek(0)
    filename = f"receipts_{year}_{month:02d}.zip"
    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Export month as Excel ────────────────────────────────────────────────────
@router.get("/export/excel")
def export_excel(
    month: int = Query(..., ge=1, le=12),
    year: int = Query(..., ge=2000),
    ctx: AuthContext = Depends(require_studio_ctx),
    repo: ExpenseRepository = Depends(get_expense_repo),
):
    from fastapi.responses import StreamingResponse
    import io
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    expenses = repo.get_multi(studio_id=ctx.studio_id, month=month, year=year, limit=1000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month:02d}-{year}"
    ws.sheet_view.rightToLeft = True

    headers = ["תאריך", "ספק", "קטגוריה", "מספר מסמך", "לפני מע\"מ", "מע\"מ", "סה\"כ", "אמצעי תשלום", "הערות", "נשלח לרו\"ח"]
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, e in enumerate(expenses, 2):
        ws.cell(row=row, column=1, value=str(e.expense_date))
        ws.cell(row=row, column=2, value=e.supplier_name or e.title)
        ws.cell(row=row, column=3, value=e.category or "")
        ws.cell(row=row, column=4, value=e.invoice_number or "")
        ws.cell(row=row, column=5, value=float(e.pretax_amount) if e.pretax_amount else "")
        ws.cell(row=row, column=6, value=float(e.vat_amount))
        ws.cell(row=row, column=7, value=float(e.amount))
        ws.cell(row=row, column=8, value=e.payment_method or "")
        ws.cell(row=row, column=9, value=e.notes or "")
        ws.cell(row=row, column=10, value="✓" if e.sent_to_accountant else "")

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"expenses_{year}_{month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Update ───────────────────────────────────────────────────────────────────
@router.put("/{expense_id}", response_model=ExpenseResponse)
def update_expense(
    expense_id: uuid.UUID,
    payload: ExpenseUpdate,
    ctx: AuthContext = Depends(require_studio_ctx),
    repo: ExpenseRepository = Depends(get_expense_repo),
):
    updated = repo.update(studio_id=ctx.studio_id, expense_id=expense_id, expense_in=payload)
    if not updated:
        raise HTTPException(status_code=404, detail="Expense not found")
    return updated


# ── Delete ───────────────────────────────────────────────────────────────────
@router.delete("/{expense_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_expense(
    expense_id: uuid.UUID,
    ctx: AuthContext = Depends(require_studio_ctx),
    repo: ExpenseRepository = Depends(get_expense_repo),
):
    exp = repo.get_by_id(studio_id=ctx.studio_id, expense_id=expense_id)
    if not exp:
        raise HTTPException(status_code=404, detail="Expense not found")
    _delete_receipt_file(exp.receipt_url, db=repo.session)
    repo.delete(studio_id=ctx.studio_id, expense_id=expense_id)
    return None


class BulkDeleteExpensesRequest(BaseModel):
    expense_ids: list[uuid.UUID]


@router.post("/bulk-delete", status_code=status.HTTP_200_OK)
def bulk_delete_expenses(
    payload: BulkDeleteExpensesRequest,
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    exps = db.query(ExpenseModel).filter(
        ExpenseModel.studio_id == ctx.studio_id,
        ExpenseModel.id.in_(payload.expense_ids),
    ).all()
    deleted = 0
    for exp in exps:
        _delete_receipt_file(exp.receipt_url, db=db)
        db.delete(exp)
        deleted += 1
    db.commit()
    return {"deleted": deleted}


# ── Attach receipt image (manual upload without OCR) ─────────────────────────
@router.post("/{expense_id}/upload-image")
async def upload_expense_image(
    expense_id: uuid.UUID,
    file: UploadFile = File(...),
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    exp = db.query(ExpenseModel).filter_by(id=expense_id, studio_id=ctx.studio_id).first()
    if not exp:
        raise HTTPException(status_code=404, detail="Expense not found")

    allowed = {"image/jpeg", "image/png", "image/webp", "image/heic", "image/heif"}
    ct = (file.content_type or "").split(";")[0].strip().lower()
    if ct not in allowed:
        raise HTTPException(status_code=400, detail="סוג קובץ לא נתמך. יש להעלות JPG, PNG או WEBP.")

    image_bytes = await file.read()
    url = _save_receipt_image(image_bytes, ct, str(ctx.studio_id), db=db)
    if not url:
        raise HTTPException(status_code=500, detail="שגיאה בשמירת התמונה")

    exp.receipt_url = url
    exp.file_size_bytes = len(image_bytes)
    db.commit()
    return {"receipt_url": url}


# ── Delete just the receipt image (keeps the expense record) ─────────────────
@router.delete("/{expense_id}/receipt-image")
def delete_expense_receipt_image(
    expense_id: uuid.UUID,
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    exp = db.query(ExpenseModel).filter_by(id=expense_id, studio_id=ctx.studio_id).first()
    if not exp:
        raise HTTPException(status_code=404, detail="Expense not found")
    if not exp.receipt_url:
        raise HTTPException(status_code=400, detail="אין תמונת קבלה למחיקה")

    _delete_receipt_file(exp.receipt_url, db=db)
    exp.receipt_url = None
    exp.file_size_bytes = None
    db.commit()
    return {"ok": True}


# ── Delete multiple receipt images at once (keeps the expense records) ───────
class BulkDeleteReceiptImagesRequest(BaseModel):
    expense_ids: list[uuid.UUID]


@router.post("/receipts/bulk-delete-images")
def bulk_delete_receipt_images(
    payload: BulkDeleteReceiptImagesRequest,
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    exps = db.query(ExpenseModel).filter(
        ExpenseModel.studio_id == ctx.studio_id,
        ExpenseModel.id.in_(payload.expense_ids),
    ).all()
    deleted = 0
    for exp in exps:
        if exp.receipt_url:
            _delete_receipt_file(exp.receipt_url, db=db)
            exp.receipt_url = None
            exp.file_size_bytes = None
            deleted += 1
    db.commit()
    return {"deleted": deleted}


# ── Storage usage summary ─────────────────────────────────────────────────────
@router.get("/storage/usage")
def get_receipt_storage_usage(
    ctx: AuthContext = Depends(require_studio_ctx),
    db: Session = Depends(get_db),
):
    from app.models.expense import Expense as ExpenseModel
    from app.models.studio import Studio
    studio = db.query(Studio).filter_by(id=ctx.studio_id).first()
    if studio:
        _reset_scan_quota_if_new_month(db, studio)

    rows = db.query(ExpenseModel).filter(
        ExpenseModel.studio_id == ctx.studio_id,
        ExpenseModel.receipt_url.isnot(None),
    ).all()

    backfilled = 0
    for exp in rows:
        if exp.file_size_bytes is None and backfilled < 25:
            size = _measure_receipt_size(exp.receipt_url)
            if size:
                exp.file_size_bytes = size
                backfilled += 1
    if backfilled:
        db.commit()

    total_bytes = sum(exp.file_size_bytes or 0 for exp in rows)
    unknown_count = sum(1 for exp in rows if exp.file_size_bytes is None)
    return {
        "total_bytes": total_bytes,
        "count": len(rows),
        "unknown_count": unknown_count,
        "scan_quota": studio.invoice_scan_quota if studio else 0,
        "scan_used": studio.invoice_scan_used if studio else 0,
        "scan_remaining": (
            max(0, studio.invoice_scan_quota - studio.invoice_scan_used)
            if studio and studio.invoice_scan_quota > 0 else None
        ),
        "scan_reset_month": studio.invoice_scan_reset_month if studio else None,
    }
